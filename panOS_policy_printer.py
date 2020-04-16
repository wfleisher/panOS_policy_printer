#!/usr/bin/env python
import re
import openpyxl
import getpass
from openpyxl.styles import Alignment, Border, PatternFill, Side
from netmiko import ConnectHandler, NetMikoAuthenticationException, NetMikoTimeoutException

## Regex Patterns
patName         = re.compile(r"([^\n\r]+)\{")
patFrom         = re.compile(r"\s{2}from\s?\[?\s?([\s\S]+?)\]?;")
patSource       = re.compile(r"\s{2}source\s\[?\s?([\s\S]+?)\]?;")
patSourceUser   = re.compile(r"\s{2}source-user\s?\[?\s?([\s\S]+?)\]?;")
patTo           = re.compile(r"\s{2}to\s?\[?\s?([\s\S]+?)\]?;")
patDestination  = re.compile(r"\s{2}destination\s?\[?\s?([\s\S]+?)\]?;")
patCategory     = re.compile(r"\s{2}category\s?\[?\s?([\s\S]+?)\]?;")
patApplication  = re.compile(r"\s{2}application\s?\[?\s?([\s\S]+?)\]?;")
patService      = re.compile(r"\s{2}service\s?\[?\s?([\s\S]+?)\]?;")
patTag          = re.compile(r"\s{2}tag\s?\[?\s?([\s\S]+?)??\]?;")
patAction       = re.compile(r"\s{2}action\s?\[?\s?([\s\S]+?)\]?;")
patDescription  = re.compile(r"\s{2}description\s[\'\"]([\s\S]+?)[\'\"];")
patDisabled     = re.compile(r"\s{2}disabled\s([\s\S]+?);")
patSpaces       = re.compile(r"(\"[A-Za-z0-9\s\S]+\"|\S+)")
patDeviceGroups = re.compile(r"Group:\s([\s\S]+?)\s")

def openSpreadheet(rowCounter):
    wb = openpyxl.Workbook()
    sheet = wb.active
    row = str(rowCounter)
    sheet["A"+row+""] = "New or Change"
    sheet["B"+row+""] = "Rule Name"
    sheet["C"+row+""] = "Source Zone"
    sheet["D"+row+""] = "Source Address"
    sheet["E"+row+""] = "User"
    sheet["F"+row+""] = "Destination zone"
    sheet["G"+row+""] = "Destination Address"
    sheet["H"+row+""] = "URL Category"
    sheet["I"+row+""] = "App-ID"
    sheet["J"+row+""] = "Service"
    sheet["K"+row+""] = "Action"
    sheet["L"+row+""] = "Description"
    sheet["M"+row+""] = "Tag"
    sheet["N"+row+""] = "Disabled"
    for col in sheet.columns:
        sheet[""+col[0].column_letter+""+row+""].fill = PatternFill(fill_type = "solid", start_color = "FF2F75B5", end_color = "FF2F75B5")
    rowCounter += 1
    return(sheet, wb, rowCounter)

def sshConnection():
    while True:
        try:
            panIP = input("Panorama IP: ")
            username = input("Username: ")
            password = getpass.getpass()
            panosConn = ConnectHandler(device_type = "paloalto_panos", host = panIP, username = username, password = password)
            break
        except (NetMikoAuthenticationException, NetMikoTimeoutException):
            print("Authentication Failed")
    panosConn.send_command("set cli scripting-mode on")
    deviceGroups = panosConn.send_command("show devicegroups | match Group")
    deviceGroups = patDeviceGroups.finditer(deviceGroups)
    deviceGroupList = []
    for match in deviceGroups:
        deviceGroupList.append(match.group(1))
    panosConn.config_mode()
    return(panosConn, deviceGroupList)

def deviceGroupGetter(rowCounter, sheet, deviceGroupList):
    deviceGroupString = ", ".join(deviceGroupList)
    deviceGroup = input("Device-group of rule ("+deviceGroupString+"): ").strip()
    while deviceGroup not in deviceGroupList:
        deviceGroup = input("Not a valid device-group. Please reenter: ")
    row = str(rowCounter)
    if sheet["A"+str(rowCounter - 1)+""].value != deviceGroup:
        sheet["A"+row+""] = deviceGroup
        for col in sheet.columns:
            sheet[""+col[0].column_letter+""+row+""].fill = PatternFill(fill_type = "solid", start_color = "FF9BC2E6", end_color = "FF9BC2E6")
        rowCounter += 1
    ruleName = ""
    return(deviceGroup, rowCounter, ruleName)

def saveFile(wb, panosConn):
    panosConn.disconnect()
    while True:
        try:
            filename = input("Filename: ")
            wb.save(filename = ""+filename+".xlsx")
            print("Policies printed to "+filename+".xlsx")
            wb.close()
            break
        except PermissionError:
            print("Permission error, close the spreadsheet")

def matchFormatter(matches):
    matchList = []
    try:
        matches = re.finditer(patSpaces, matches.group(1))
    except TypeError:
        return(matchList)
    for match in matches:
        match = match.group(1)
        matchList.append(match)
    matchList = "\n".join(matchList).replace("\"","")
    return(matchList)

def policyPrinter(rowCounter, sheet, deviceGroup, ruleName, panosConn):
    policy = panosConn.send_command("show device-group "+deviceGroup+" post-rulebase security rules \""+ruleName+"\"")
    try:
        policyName = patName.search(policy).group(1).replace("\"","").strip()
    except AttributeError:
        return(rowCounter, ruleName)
    policyFrom = matchFormatter(patFrom.search(policy))
    policySource = matchFormatter(patSource.search(policy))
    policySourceUser = matchFormatter(patSourceUser.search(policy))
    policyTo = matchFormatter(patTo.search(policy))
    policyDestination = matchFormatter(patDestination.search(policy))
    policyCategory = matchFormatter(patCategory.search(policy))
    policyApplicaation = matchFormatter(patApplication.search(policy))
    policyService = matchFormatter(patService.search(policy))
    policyAction = patAction.search(policy).group(1)
    try:
        policyDescription = patDescription.search(policy).group(1).strip()
    except AttributeError:
        policyDescription = ""
    try:
        policyTag = matchFormatter(patTag.search(policy))
        if policyTag == []:
            policyTag = ""
    except AttributeError:
        policyTag = ""
    try:
        policyDisabled = patDisabled.search(policy).group(1)
    except AttributeError:
        policyDisabled = "no"
    row = str(rowCounter)
    sheet["B"+row+""] = policyName             ##name
    sheet["C"+row+""] = policyFrom             ##src zone
    sheet["D"+row+""] = policySource           ##src address
    sheet["E"+row+""] = policySourceUser       ##src user
    sheet["F"+row+""] = policyTo               ##dst zone
    sheet["G"+row+""] = policyDestination      ##dst address
    sheet["H"+row+""] = policyCategory         ##URL
    sheet["I"+row+""] = policyApplicaation     ##app
    sheet["J"+row+""] = policyService          ##service
    sheet["K"+row+""] = policyAction           ##action
    sheet["L"+row+""] = policyDescription      ##description
    sheet["M"+row+""] = policyTag              ##tag
    sheet["N"+row+""] = policyDisabled         ##Disabled
    rowCounter += 1
    ruleName = ""
    return(rowCounter, ruleName)

def xlFormater(rowCounter, sheet):
    for row in sheet.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal = "center", vertical = "bottom",wrapText = True)
            cell.border = Border(
                left = Side(border_style = "thin", color = "FF000000"),
                right = Side(border_style = "thin", color = "FF000000"),
                top = Side(border_style = "thin", color = "FF000000"),
                bottom = Side(border_style = "thin", color = "FF000000"))
    for col in sheet.columns:
        length = 0
        for cell in col:
            if cell.value != None:
                if "\n" in cell.value:
                    cellValueList = cell.value.split("\n")
                    cellLength = (max(len(cellValue) for cellValue in cellValueList) + 3)
                    if cellLength > length:
                        length = cellLength
                elif (len(cell.value) + 3) > length:
                    length = (len(cell.value) + 3)
        if length > 85:
            sheet.column_dimensions[col[0].column_letter].width = 85
        else:
            sheet.column_dimensions[col[0].column_letter].width = length

def main():
    rowCounter = 1
    sheet, wb, rowCounter = openSpreadheet(rowCounter)
    panosConn, deviceGroupList = sshConnection()
    deviceGroup, rowCounter, ruleName = deviceGroupGetter(rowCounter, sheet, deviceGroupList)
    while True:
        if ruleName == "":
            ruleName = input("Rule name (Type 'change' to switch device-group or 'quit' to finish): ").strip()
        elif ruleName == "change":
            deviceGroup, rowCounter, ruleName = deviceGroupGetter(rowCounter, sheet, deviceGroupList)
            continue
        elif ruleName == "quit":
            xlFormater(rowCounter, sheet)
            saveFile(wb, panosConn)
            break
        else:
            rowCounter, ruleName = policyPrinter(rowCounter, sheet, deviceGroup, ruleName, panosConn)
            if ruleName == "change" or ruleName == "quit":
                continue
            elif ruleName != "":
                ruleName = input("Not a valid rule name. Please reenter: ")

main()